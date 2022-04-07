"""
=============================
author:'liudl'
time:2019/8/21
E-mail:liudl7@lenovo.com
=============================
"""
import openpyxl


class CaseData(object):
    """测试用例数据类"""

    def __init__(self, zip_obj, *args, **kwargs):
        # 变量zip对象
        for i in list(zip_obj):
            # 通过反射机制设置属性
            # 将表头作为属性，值作为属性值
            setattr(self, i[0], i[1])


# 定义一个类专门用来读取Excel中的数据

class ReadExcel(object):
    """读取Excel中的用例数据,读取出来的是一个字典"""

    def __init__(self, file_name, sheet_name):
        """

        :param file_name: excel文件名
        :param sheet_name:sheet表单名
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        """打开工作簿和表单"""
        # 打开文件，返回一个工作簿对象
        self.workbook = openpyxl.load_workbook(self.file_name)
        # 通过工作簿，选择表单对象
        self.sheet = self.workbook[self.sheet_name]

    def read_data(self):
        """读取所有用例数据
            读取出来的是一个列表，列表中存的是字典
        """

        # 打开文件和表单
        self.open()
        # 按行获取所有的表格对象，每一行内容放在一个元组中,并转换为列表
        rows = list(self.sheet.rows)
        # 创建一个列表cases，存放所有的用例数据
        cases = []
        # 获取表头
        titles1 = [row.value for row in rows[0]]
        # 遍历其它的数据行，和表头进行打包，转换为字典，放到cases这个列表中
        for row in rows[1:]:
            data = [r.value for r in row]
            # 和表头进行打包，转换为字典
            case = dict(zip(titles1, data))
            cases.append(case)
        # 将读取出来的数据进行返回
        return cases

    def read_data_obj(self):
        """

        :return: 读取出来的是一个列表，列表中存的是对象
        """
        # 打开工作簿
        self.open()
        # 创建一个空的列表，用例存放所有的用例数据
        cases = []
        # 读取表单中的数据
        rows = list(self.sheet.rows)
        # 读取表头
        titles = []
        for r in rows[0]:
            titles.append(r.value)

        # 按行遍历
        for row in rows[1:]:
            # 读每一行数据
            case = []
            for r in row:
                case.append(r.value)
            zip_obj = zip(titles, case)
            # 将每一条用例的数据，存储为一个对象
            # 通过Case这个类来创建一个对象，传了一个参数，zip_obj
            case_data = CaseData(zip_obj)
            cases.append(case_data)
        # 将包含所有用例的列表cases进从返回
        return cases

    # def read_data_dict2(self):
    #     self.open()
    #     cases = []
    #     max_row = self.sh.max_row
    #     for r in range(1, max_row + 1):
    #         case_dict = []
    #         case_dict['case_id'] = self.sh.cell(r, 1).value
    #         case_dict['excepted'] = self.sh.cell(r, 2).value
    #         case_dict['data'] = self.sh.cell(r, 3).value
    #         cases.append(cases)

    def write_data(self, row, column, value):
        """
        这是写入文件的方法
        :param row: 写入的行
        :param column: 写入的列
        :param value:写入的内容
        :return:
        """
        # 打开文件
        self.open()
        # 按照传入的行、列、内容进行写入
        self.sheet.cell(row=row, column=column, value=value)
        # 保存
        self.workbook.save(self.file_name)


# if __name__ == '__main__':
#     do_excel = ReadExcel('cases.xlsx', 'Sheet1')
#     do_excel.read_data_obj()
#     cases = do_excel.read_data_obj()

    # cases = do_excel.read_data()
    # print(cases)
